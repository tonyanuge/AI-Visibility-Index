"""Import captures from the AI Mystery Shopping tracker (xlsx Run Log) or a CSV."""
import csv
from ..core.models import Mention
from ..store import repository as repo

_COLS = ["Date","Area","Category","Engine","Archetype","Prompt Text",
         "Business Named","Rank","Accuracy","Sentiment","Source"]

def _to_mention(row: dict) -> Mention:
    rank = row.get("Rank")
    try:
        rank = int(rank) if rank not in (None, "", "-") else None
    except (TypeError, ValueError):
        rank = None
    return Mention(
        business=(row.get("Business Named") or "").strip(),
        category=(row.get("Category") or "").strip(),
        area=(row.get("Area") or "").strip(),
        engine=(row.get("Engine") or "").strip(),
        archetype=(row.get("Archetype") or "Discovery").strip(),
        rank=rank,
        accuracy=(row.get("Accuracy") or "N/A").strip(),
        sentiment=(row.get("Sentiment") or "Neutral").strip(),
        source=(row.get("Source") or row.get("Source Cited") or "Unclear").strip(),
        prompt_text=(row.get("Prompt Text") or "").strip(),
        date=(row.get("Date") or "").strip(),
    )

def import_csv(conn, path) -> int:
    n = 0
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m = _to_mention(row)
            if m.business:
                repo.add_mention(conn, m)
                repo.add_business(conn, m.business, m.category, m.area)
                n += 1
    return n

def import_tracker_xlsx(conn, path, sheet="Run Log", header_row=3) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in ws[header_row]]
    n = 0
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[i]: r[i] for i in range(len(headers)) if headers[i]}
        if not row.get("Business Named") or "EXAMPLE" in str(row.get("Business Named")):
            continue
        m = _to_mention({k: ("" if v is None else str(v)) for k, v in row.items()})
        if m.business:
            repo.add_mention(conn, m)
            repo.add_business(conn, m.business, m.category, m.area)
            n += 1
    return n
